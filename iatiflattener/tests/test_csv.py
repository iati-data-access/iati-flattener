import os
import pytest
import iatiflattener
from iatiflattener import model
from iatiflattener import group_data
import csv
import shutil
import openpyxl


class TestModel():

    @pytest.fixture(scope='class')
    def flattener(self):
        try:
            shutil.rmtree('output_test/')
        except FileNotFoundError:
            pass
        yield iatiflattener.FlattenIATIData(
            refresh_rates=False,
            iatikitcache_dir='',
            output='output_test',
            publishers=None,
            langs=['en', 'fr'],
            run_publishers=False,
            exchange_rates_filename="iatiflattener/tests/fixtures/rates.csv"
        )


    @pytest.fixture(scope='class')
    def package(self, flattener):
        flattener.process_package(None, f"fcdo-activity.xml",
            'iatiflattener/tests/fixtures')


    @pytest.fixture()
    def csvfile(self, package):
        with open('output_test/csv/transaction-LR.csv', 'r') as _csv_file:
            yield csv.DictReader(_csv_file)


    @pytest.fixture()
    def activity_csvfile(self, package):
        with open('output_test/csv/activities/GB-GOV-1.csv', 'r') as _csv_file:
            yield csv.DictReader(_csv_file)


    @pytest.fixture(scope='class')
    def group(self, package):
        group_data.GroupFlatIATIData(
            langs=['en', 'fr'],
            output_folder='output_test'
            )


    @pytest.fixture
    def groupedfile_en(self, package, group):
        yield openpyxl.load_workbook(filename='output_test/xlsx/en/LR.xlsx')


    @pytest.fixture
    def groupedfile_fr(self, package, group):
        yield openpyxl.load_workbook(filename='output_test/xlsx/fr/LR.xlsx')


    def test_activity_row(self, flattener, package, activity_csvfile):
        first_row = next(activity_csvfile)
        assert first_row == {
            'iati_identifier': 'GB-1-103662-101',
            'title#en': 'PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia',
            'title#fr': 'PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia',
            'description#en': "This activity (PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia) is a component of Civil Ser. Cap. Bldng. Liberia reported by FCDO, with a funding type of 'Procurement of Services' and a budget of £2,721,510. This component benefits Liberia, and works in the following sector(s): Public sector policy and administrative management. , with the following implementing partners: Adam Smith International. The start date is 01-12-2006 and the end date is 31-03-2010.",
            'description#fr': "This activity (PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia) is a component of Civil Ser. Cap. Bldng. Liberia reported by FCDO, with a funding type of 'Procurement of Services' and a budget of £2,721,510. This component benefits Liberia, and works in the following sector(s): Public sector policy and administrative management. , with the following implementing partners: Adam Smith International. The start date is 01-12-2006 and the end date is 31-03-2010.",
            'reporting_org_group': 'GB',
            'reporting_org#en': 'UK - Foreign, Commonwealth and Development Office [GB-GOV-1]',
            'reporting_org#fr': 'Royaume-Uni – Ministère des Affaires étrangères, du Commonwealth et du Développement [GB-GOV-1]',
            'reporting_org_ref': 'GB-GOV-1',
            'location': '',
            'start_date': '2006-12-01',
            'end_date': '2010-03-31',
            'GLIDE': '',
            'HRP': '',
            'hash': '71c4ddbc047173a4a9a1c10635cc03d9'
        }


    def test_row(self, flattener, package, csvfile):
        row = next(csvfile)
        assert row == {
            'iati_identifier': 'GB-1-103662-101',
            'title#en': 'PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia',
            'title#fr': 'PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia',
            'reporting_org_group': 'GB',
            'reporting_org#en': 'UK - Foreign, Commonwealth and Development Office [GB-GOV-1]',
            'reporting_org#fr': 'Royaume-Uni – Ministère des Affaires étrangères, du Commonwealth et du Développement [GB-GOV-1]',
            'reporting_org_type': '10', 'aid_type': 'C01', 'finance_type': '110',
            'flow_type': '10',
            'provider_org#en': 'UK - Foreign, Commonwealth and Development Office [GB-GOV-1]',
            'provider_org#fr': 'Royaume-Uni – Ministère des Affaires étrangères, du Commonwealth et du Développement [GB-GOV-1]',
            'provider_org_type': '', 'receiver_org#en': 'Adam Smith International [GB-COH-2732176]',
            'receiver_org#fr': 'Adam Smith International [GB-COH-2732176]', 'receiver_org_type': '70',
            'transaction_type': '3', 'value_original': '1232.0', 'currency_original': 'GBP',
            'value_usd': '1696.3408008150918', 'exchange_rate_date': '2021-08-31',
            'exchange_rate': '0.726269155', 'value_eur': '1433.446680400464',
            'value_local': '0.0',
            'transaction_date': '2010-05-27', 'country_code': 'LR', 'multi_country': '0',
            'sector_category': '150', 'sector_code': '15110', 'humanitarian': '0', 'fiscal_year': '2010',
            'fiscal_quarter': 'Q2',
            'fiscal_year_quarter': '2010 Q2', 'url': 'https://d-portal.org/q.html?aid=GB-1-103662-101'
        }


    def test_group_headers_en(self, flattener, package, groupedfile_en):
        headers_row = [cell.value for cell in next(groupedfile_en.worksheets[0].iter_rows(min_row=1, max_row=1))]
        assert headers_row == ['IATI Identifier', 'Title', 'Reporting Organisation Group', 'Reporting Organisation', 'Reporting Organisation Type', 'Aid Type', 'Finance Type', 'Flow Type', 'Provider Organisation', 'Provider Organisation Type', 'Receiver Organisation', 'Receiver Organisation Type', 'Transaction Type', 'Recipient Country or Region', 'Multi Country', 'Sector Category', 'Sector', 'Humanitarian', 'Calendar Year', 'Calendar Quarter', 'Calendar Year and Quarter', 'URL', 'Value (USD)', 'Value (EUR)', 'Value (Local currrency)']

    def test_first_row_en(self, flattener, package, groupedfile_en):
        first_row = [cell.value for cell in next(groupedfile_en.worksheets[0].iter_rows(min_row=2, max_row=2))]
        assert first_row == ['GB-1-103662-101', 'PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia', 'GB - United Kingdom', 'UK - Foreign, Commonwealth and Development Office [GB-GOV-1]', '10 - Government', 'C01 - Project-type interventions', '110 - Standard grant', '10 - ODA', 'UK - Foreign, Commonwealth and Development Office [GB-GOV-1]', 'No data', 'Adam Smith International [GB-COH-2732176]', '70 - Private Sector', '3 - Disbursement', 'LR - Liberia', 0, '150 - Government & Civil Society', '15110 - Public sector policy and administrative management', 0, 2010, 'Q2', '2010 Q2', 'https://d-portal.org/q.html?aid=GB-1-103662-101', 1696.34080081509, 1433.44668040046, 0]

    def test_group_headers_fr(self, flattener, package, groupedfile_fr):
        headers_row = [cell.value for cell in next(groupedfile_fr.worksheets[0].iter_rows(min_row=1, max_row=1))]
        assert headers_row == ['Identifiant de l’IITA', 'Titre', 'Groupe d’organisme déclarant', 'Organisme déclarant', 'Type d’organisme déclarant', 'Type d’aide', 'Type de financement', 'Type de flux', 'Organisme prestataire', 'Type d’organisme prestataire', 'Organisme bénéficiaire', 'Type d’organisme bénéficiaire', 'Type de transaction', 'Pays ou région bénéficiaire', 'Multipays', 'Catégorie de secteur', 'Secteur', 'Humanitaire', 'Année civile', 'Trimestre civil', 'Année et trimestre civils', 'URL', 'Valeur (USD)', 'Valeur (EUR)', 'Valeur (Monnaie locale)']

    def test_first_row_fr(self, flattener, package, groupedfile_fr):
        first_row = [cell.value for cell in next(groupedfile_fr.worksheets[0].iter_rows(min_row=2, max_row=2))]
        assert first_row == ['GB-1-103662-101', 'PROCOFSERVICES and P0220 for Civil Ser. Cap. Bldng. Liberia', 'GB - Royaume-Uni (le)', 'Royaume-Uni – Ministère des Affaires étrangères, du Commonwealth et du ' 'Développement [GB-GOV-1]', '10 - Gouvernement', 'C01 - Interventions de type projet', '110 - Dons ordinaires', '10 - APD', 'Royaume-Uni – Ministère des Affaires étrangères, du Commonwealth et du ' 'Développement [GB-GOV-1]', 'Aucune donnée', 'Adam Smith International [GB-COH-2732176]', '70 - Secteur privé', '3 - Décaissement', 'LR - Libéria (le)', 0, '150 - Gouvernement & Société Civile', '15110 - Politiques publiques et gestion administrative', 0, 2010, 'Q2', '2010 Q2', 'https://d-portal.org/q.html?aid=GB-1-103662-101', 1696.34080081509, 1433.44668040046, 0]
